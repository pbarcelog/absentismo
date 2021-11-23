import openpyxl
from itertools import islice
import unidecode

def build_dictionary(xlsx_dictionary):
    dictionary = {'sexo': {'Masculino': 'masculino', 'Femenino': 'femenino'}, \
                  'clase_de_contrato': {'Indefinido': 'indefinido', 'Temporal': 'temporal',
                                        'Contrato temporal': 'temporal'}, \
                  'posicion': {'AYUDANTE/A DE SISTEMAS': 'AYUDANTE/A SISTEMAS', \
                               'TECNICO/A ADMINISTRATIVO': 'TECNICO/A ADMINISTRATIVO/A', \
                               'TECNICO/A DE SISTEMAS': 'TECNICO/A SISTEMAS'}, \
                  'identificacion': {'Indefinido': 'indefinido', 'Contrato temporal': 'temporal', 'Temporal': 'temporal', \
                                     'Contrato ordinario por tiempo indefinido': 'Indefinido Tiempo completo', \
                                     'Indefinido.Tiempo completo': 'Indefinido Tiempo completo'}, \
                  'descripcion':  {'Física Otra Clasificación': 'Fisica', 'Física Otra clasificación': 'Física',\
                                    'Afiliado a Once Con Resto Visual': 'Afiliados Con Resto Visual', \
                                    'Afiliado a Once Sin Resto Visual': 'Afiliados Sin Resto Visual'}
                  }

    # abrir el excel que contiene los diccionarios de orden, centro y especialidad

    wb_dictionary = openpyxl.load_workbook(xlsx_dictionary)

    # lista con las keys del diccionario inicial de categorías
    list_excel_categorias = list(dictionary.keys())

    # abrir el excel de los diccionarios de categorías
    for name in wb_dictionary.sheetnames:
        # añadir al diccionario inicial de categorías las que vengan de los diccionarios en el excel
        list_excel_categorias.append(name)
        # abrir el excel hoja a hoja
        parameter = wb_dictionary[name]
        # crear diccionario temporal
        parameter_dict = {}
        # construir el diccionario de una hoja (columna izquierda se traduce en columna derecha)
        for row in islice(parameter.values, 0, parameter.max_row):
            parameter_dict[row[0]] = row[1]
        # introducir nuevo diccionario en el diccionario de diccionarios (nombre hoja excel = diccionario de esa hoja)
        dictionary[name] = parameter_dict

    wb_dictionary.close()

    return dictionary, list_excel_categorias

def headers(wb, list_excel_categorias):

    sheet = wb.active

    # generar los headers que se usarán en todas las lecturas: eliminar problema de diferencias de headers
    header_original = []
    list_header_categorias = []
    dict_worker = {}
    # coger solo la primera fila, la de los headers
    for value in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for header in value:
            # tratar header para que no tenga ni acentos ni puntos ni mayúsculas ni espacios
            header_final = unidecode.unidecode(header).replace('.', '').lower().replace(' ', '_')
            # print(header, ' ', header_final)
            # añadir el header tratado a la lista de headers
            # print(header_final in header_original)
            if header_final in header_original:
                header_final = header_final + '_hijo'
                # print(header_final)
            header_original.append(header_final)
            # añadir el header a la lista de categorías que se han de traducir de header a diccionario de diccionarios
            for categoria in list_excel_categorias:
                if categoria in header_final:
                    list_header_categorias.append(header_final)

    return header_original, list_header_categorias