import references
import create
import openpyxl
from pathlib import Path
import pandas as pd
from datetime import datetime, timedelta
import json
from collections import OrderedDict
from itertools import islice
import unidecode
import numpy as np

import os
import sys
#import findspark ######### PROBAR SI SE ESTRELLA ##########
#findspark.init() #########      CON TAREAS       ##########
import pyspark
import pyspark.sql
# from pyspark.shell import sqlContext
from pyspark.sql import Row, SparkSession
from pyspark.sql.types import *
import pyspark.sql.functions as F
from pyspark.sql.functions import col, lit, when

# from pyspark.sql import Row

HADOOP_HOME = "C:\\hadoop"

os.environ["HADOOP_HOME"] = HADOOP_HOME
sys.path.append(HADOOP_HOME + "\\bin")
os.environ['PYSPARK_PYTHON'] = sys.executable
os.environ['PYSPARK_DRIVER_PYTHON'] = sys.executable

def workers(output_data, excels, num_cpu):

    ###################################################
    ####   CARGA DE DATOS WORKERS Y SESION SPARK   ####
    ####         DROP DE VARIAS COLUMNAS           ####
    ###################################################

    now = datetime.now()
    now_inicial = now

    # iniciar el diccionario de categorías con las conversiones más obvias
    dictionary, list_excel_categorias = references.build_dictionary(Path(Path.home(), 'Final RR', 'diccionario.xlsx'))
    print(dictionary)
    print(list_excel_categorias)

    # abrir archivo excel con hojas excel con info trabajadores de 2014 a 2020
    xlsx_worker = Path(Path.home(), 'Final RR', 'CC DATOS TRABAJADORES 2014 - 2020 HIJOS' + excels)
    wb_worker = openpyxl.load_workbook(xlsx_worker)
    header_original, list_header_categorias = references.headers(wb_worker, list_excel_categorias)

    print('\nHeaders - Categorías: ', list_header_categorias)
    print('\nHeaders Originales: ', header_original)

    # equivalencias entre headers y keys del diccionario
    names_dict = {}
    for categoria in list_excel_categorias:
        for cual in list_header_categorias:
            if categoria in cual:
                names_dict[cual] = categoria

    print(names_dict)

    # names_dict = dict(zip(list_excel_categorias, list_header_categorias))
    # print('Names - Dict: ', names_dict)
    print('Headers y diccionarios completados en: ', datetime.now() - now)

    now = datetime.now()

    df = pd.DataFrame(columns=header_original)
    for name in wb_worker.sheetnames:  # cojemos sheets del fichero una a una (2014 a 2020)
        # print(i)
        print(name)
        sheet = wb_worker[name]  # cargar la sheet correspondiente
        new_data = [create.create_row_worker(row, header_original, names_dict, dictionary)
                    for row in islice(sheet.values, 1, sheet.max_row)]
        # print(new_data)
        df = df.append(new_data, ignore_index=True)


    wb_worker.close()
    ### columnas no relevantes o que no conviene tener ###
    KEY_DROP = ['sociedad', 'edad_del_empleado', 'posicion', 'especialidad', 'sigla_de_via_publica', 'calle_y_numero',
                'nocasa', 'vivienda', 'poblacion', 'telefono', 'numero_de_personal', 'nif', 'fecha_nac_hijo',
                'clave_de_sexo_hijo', 'noafil']

    df = df.drop(KEY_DROP, axis=1).drop_duplicates()

    print('\nDF Workers en: ', datetime.now() - now)

    now = datetime.now()

    # sesión spark con conector a bd mongo donde se salvarán los datos procesados. Atención a la memoria.
    spark = SparkSession.builder.master("local["+num_cpu+"]").appName("worker") \
        .config("spark.mongodb.output.uri", "mongodb://localhost:27017/" + output_data)\
        .config("spark.jars.packages", "org.mongodb.spark:mongo-spark-connector_2.12:3.0.1") \
        .config("spark.executor.memory", "3g") \
        .config("spark.driver.memory", "8g") \
        .getOrCreate()

    print('\nSpark Session completada en: ', datetime.now() - now)

    now = datetime.now()

    ############################################
    ### DE PANDAS A SPARK; CASTING COLUMNAS; ###
    ### DATAFRAME TEMPORAL PARA AGREGACIÓN;  ###
    ###    PRIMER FICHERO DE CONTROL CSV     ###
    ############################################

    KEY = ['division_de_personal', 'no_pers', 'clave_de_sexo', 'descripcion', 'porcentaje_de_discapacidad',
           'numero_de_orden', 'centro_de_coste', 'categoria_profesional', 'clase_de_contrato',
           'identificacion_de_contratos_se',
           'h_sem', 'feantig', 'desde', 'hasta', 'motivo_de_la_medida', 'estado_civil', 'fecha_nac', 'nacionalidad',
           'codpostal', 'clase_de_instituto_de_ensenanz', 'clase_registro_parentesco', 'idobj']

    structureSchema = StructType([
        StructField('division_de_personal', StringType(), True), StructField('no_pers', StringType(), True),
        StructField('clave_de_sexo', StringType(), True), StructField('descripcion', StringType(), True),
        StructField('porcentaje_de_discapacidad', StringType(), True),
        StructField('numero_de_orden', StringType(), True),
        StructField('centro_de_coste', StringType(), True), StructField('categoria_profesional', StringType(), True),
        StructField('clase_de_contrato', StringType(), True),
        StructField('identificacion_de_contratos_se', StringType(), True),
        StructField('h_sem', FloatType(), True), StructField('feantig', StringType(), True),
        StructField('desde', StringType(), True), StructField('hasta', StringType(), True),
        StructField('motivo_de_la_medida', StringType(), True), StructField('estado_civil', StringType(), True),
        StructField('fecha_nac', StringType(), True), StructField('nacionalidad', StringType(), True),
        StructField('codpostal', StringType(), True), StructField('clase_de_instituto_de_ensenanz', StringType(), True),
        StructField('idobj', StringType(), True), StructField('clase_registro_parentesco', StringType(), True)
    ])

    sparkDF = spark.createDataFrame(df, schema=structureSchema)
    print('\nPasar de df a sparkdf completado en: ', datetime.now() - now)
    now = datetime.now()

    for col in ['feantig', 'desde', 'hasta', 'fecha_nac']:
        sparkDF = sparkDF.withColumn(col, sparkDF[col].cast(DateType()))

    for col in ['idobj', 'codpostal', 'no_pers']:
        sparkDF = sparkDF.withColumn(col, sparkDF[col].cast(IntegerType()))

    KEY = ['division_de_personal', 'no_pers', 'clave_de_sexo', 'descripcion', 'porcentaje_de_discapacidad',
           'numero_de_orden', 'centro_de_coste', 'categoria_profesional', 'clase_de_contrato',
           'identificacion_de_contratos_se', 'h_sem', 'feantig', 'desde', 'hasta', 'motivo_de_la_medida', 'fecha_nac',
           'nacionalidad', 'codpostal', 'clase_de_instituto_de_ensenanz', 'clase_registro_parentesco']

    sparkDF = sparkDF.withColumn('porcentaje_de_discapacidad',
                                 sparkDF.porcentaje_de_discapacidad.cast(FloatType()))\
                    .withColumn('porcentaje_de_discapacidad', F.round(F.col('porcentaje_de_discapacidad'), 2))\
                    .withColumn('h_sem', F.round(F.col('h_sem'), 1))\
                    .withColumn("division_de_personal", when(F.col('division_de_personal') == '', F.lit('desconocida'))
                                .otherwise(F.col('division_de_personal'))) \
                    .withColumn('idobj',
                                 when(F.col('clase_registro_parentesco') != "", F.lit(1))
                                .when((F.col('clase_registro_parentesco') == "") & (F.col('idobj').isNotNull()) &
                                      (F.col('idobj') != 0), F.lit(1))
                                .otherwise(F.lit(0)))\
                    .withColumn('clave_de_sexo', when(F.col('clave_de_sexo') == '', F.lit('indeterminado'))
                                                .otherwise(F.col('clave_de_sexo')))\
                    .withColumn('clase_registro_parentesco', when((F.col('clase_registro_parentesco') == "") &
                                                                  (F.col('idobj') == 1), F.lit('alguno'))
                                                            .when((F.col('clase_registro_parentesco') == "") &
                                                                  (F.col('idobj') == 0), F.lit('ninguno'))
                                                            .otherwise(F.col('clase_registro_parentesco')))\
                    .withColumn('numero_de_orden',
                                 when(F.col('numero_de_orden') != "", F.col('numero_de_orden'))
                                 .otherwise(F.lit('Multiservicio')))\
                    .withColumn('codpostal', when(F.col('codpostal').isNull(), F.lit('00000')).otherwise(F.col('codpostal')))\
                    .withColumn('feantig', when(F.col('feantig').isNull(), F.col('desde')).otherwise(F.col('feantig')))\
                    .withColumn('centro_de_coste', when(F.col('centro_de_coste') == "", F.lit('indefinido'))
                                                    .otherwise(F.col('centro_de_coste'))) \
                    .withColumn('motivo_de_la_medida', when(F.col('motivo_de_la_medida') == "", F.lit('indeterminada'))
                                                    .otherwise(F.col('motivo_de_la_medida'))) \
                    .withColumn('categoria_profesional', when(F.col('categoria_profesional') == "", F.lit('indefinida'))
                                                    .otherwise(F.col('categoria_profesional'))) \
                    .withColumn('clase_de_contrato', when(F.col('clase_de_contrato') == "", F.lit('indeterminado'))
                                                    .otherwise(F.col('clase_de_contrato'))) \
                    .withColumn('nacionalidad', when(F.col('nacionalidad') == "", F.lit('indeterminada'))
                                                .otherwise(F.col('nacionalidad'))) \
                    .withColumn('identificacion_de_contratos_se', when(F.col('identificacion_de_contratos_se') == "",
                                                                        F.lit('desconocida'))
                                                                    .otherwise(F.col('identificacion_de_contratos_se')))\
                    .groupby(KEY).sum('idobj').withColumnRenamed("sum(idobj)", "idobj")\
                    .dropDuplicates()

    # sparkDF.toPandas().to_csv('c:\\tmp\\workers.csv')
    #a = input('check workers en tmp')

    #a = input('check workers.csv')

    sparkDF.printSchema()
    print('\nFilter completado en: ', datetime.now() - now)


    ##############################################################
    ###   GROUPBY PARA DETERMINAR FECHA MÁXIMA DE INICIO QUE   ###
    ###       PUEDE TENER EL ÚLTIMO CONTRATO VIGENTE           ###
    ##############################################################

    from pyspark.sql.functions import max as max_

    now = datetime.now()

    sparkDF.createOrReplaceTempView("datos_trabajador")
    sparkDF_tmp = sparkDF.filter('hasta <= current_date()')

    KEY = ['division_de_personal', 'no_pers', 'clave_de_sexo', 'descripcion', 'porcentaje_de_discapacidad',
           'numero_de_orden', 'centro_de_coste', 'categoria_profesional', 'clase_de_contrato',
           'identificacion_de_contratos_se', 'h_sem', 'feantig', 'fecha_nac', 'nacionalidad',
           'codpostal', 'clase_de_instituto_de_ensenanz', 'clase_registro_parentesco', 'idobj']

    sparkDF_tmp2 = sparkDF_tmp.withColumn('datetime', F.col('hasta').cast('timestamp')).groupby(KEY).agg(max_('hasta'))
    sparkDF_tmp2 = sparkDF_tmp2.withColumnRenamed("max(hasta)", "hasta_max")\
        .withColumn('hasta_max', F.date_add(F.col('hasta_max'), 1))
    sparkDF_tmp.unpersist()

    print('\nPasar max de hasta en: ', datetime.now() - now)

    ###############################################################
    ###      JOIN DE DF TEMPORAL CON SPARKDF PARA INCLUIR       ###
    ###   COLUMNA CON LA FECHA MÁXIMA DEL PENÚLTIMO CONTRATO    ###
    ###############################################################

    sparkDF.createOrReplaceTempView("datos_trabajador")
    sparkDF_tmp2.createOrReplaceTempView("max")

    sql = "SELECT a.*, b.hasta_max FROM datos_trabajador AS a LEFT JOIN max as b " \
          "ON b.no_pers = a.no_pers"

    KEY.remove('no_pers')
    adition = ""
    for element in KEY:
        adition += " and b." + element + " = a." + element
    sql += adition
    print(sql)

    sparkDF = spark.sql(sql)
    sparkDF_tmp2.unpersist()
    #sparkDF.toPandas().to_csv('c:\\tmp\\workers2.csv')

    #a = input('check workers2')

    ###############################################################
    ###    CORRECCIÓN DE FECHAS: SI ES INDEFINIDO, LA FECHA     ###
    ### LÍMITE ES FINAL DE AGOSTO; SI HAY OVERLAP, SE COGE COMO ###
    ###  FECHA DE INICIO LA FECHA FINAL DEL CONTRATO ANTERIOR   ###
    ###                 SEGUNDO CSV DE CONTROL                  ###
    ###############################################################

    low_bound = '2014-01-01'
    high_bound = '2020-12-31'
    low_bound = F.to_date(F.lit(low_bound))
    high_bound = F.to_date(F.lit(high_bound))
    sparkDF.createOrReplaceTempView("datos_trabajador")
    # sparkDF.printSchema()
    sparkDF = sparkDF.withColumn("year", F.year(F.col('hasta'))) \
        .withColumn("desde_real", when((F.col('year') == 9999) & (F.col('hasta_max') != ""),
                                       F.col('hasta_max')) \
                    .when(F.col('desde') < low_bound, low_bound) \
                    .otherwise(F.col('desde'))) \
        .withColumn("hasta_real", when(F.col('hasta') > high_bound, high_bound) \
                    .otherwise(F.col('hasta'))) \
        .withColumn('dias_rango', F.datediff(F.date_add(F.col('hasta_real'), 1), F.col('desde_real')))

    ################################################################
    ###  GROUPBY PARA CONTAR EL NÚMERO DE PROYECTOS EN LOS QUE   ###
    ###    UN TRABAJADOR PUEDE ESTÁ TRABAJANDO APARENTEMENTE     ###
    ### SIMULTÁNEAMENTE EN EL PERIODO DE VIGENCIA DE UN CONTRATO ###
    ################################################################

    now = datetime.now()

    # lista de todos los parámetros menos la división, el centro de coste y el número de orden (proyecto/servicio)
    # en el que trabaja cada empleado
    key = ['no_pers', 'clave_de_sexo', 'descripcion', 'porcentaje_de_discapacidad',
           'categoria_profesional', 'clase_de_contrato', 'identificacion_de_contratos_se',
           'h_sem', 'feantig', 'motivo_de_la_medida', 'fecha_nac', 'nacionalidad', 'codpostal',
           'clase_de_instituto_de_ensenanz', 'clase_registro_parentesco', 'idobj', 'desde_real', 'hasta_real']

    # bucle para construir las condiciones de un join con todos los parámetros menos el número de orden
    condition = 'a.no_pers == b.no_pers'
    for element in key[1:]:
        condition = condition + ' and ' + 'a.' + element + ' == ' + 'b.' + element
    print(condition)

    # contamos en cuántos proyectos/servicios está presente un empleado simulatáneamente durante la vigencia de un contrato
    b = sparkDF.groupBy(key).count()
    # crear dos vistas temporales, de los trabajadores y de la agregación para el número de proyectos
    sparkDF.createOrReplaceTempView("datos_trabajador")
    b.createOrReplaceTempView("num_proyectos")
    # hacer un join para juntar cada empleado con el número de proyectos/servicios en los que está presente
    sql = "SELECT a.*, b.count AS proyectos FROM datos_trabajador AS a LEFT JOIN num_proyectos as b ON " + condition
    sparkDF = spark.sql(sql)
    # sparkDF.printSchema()
    # sparkDF.show()
    # sparkDF.toPandas().to_csv('c:\\tmp\\workers3.csv')
    # a = input('check workers3: ')
    b.unpersist()

    ###############################################################
    ###   SE IMPORTAN LAS AUSENCIAS; DROP DE ALGUNAS COLUMNAS   ###
    ### SE GENERA DF SPARK; CAST DE COLUMNAS; FILTER DE GUARDAS ###
    ###############################################################

    now = datetime.now()

    xlsx_absence = Path(Path.home(), 'Final RR', 'CC CONSULTA ABSENTISMOS 2014 - 2020' + excels)
    wb_absence = openpyxl.load_workbook(xlsx_absence)
    header_absence, dummy_list = references.headers(wb_absence, list_excel_categorias)

    #columns = header_absence + ['intervalo']
    #print(columns)
    df_absence = pd.DataFrame(columns=header_absence)

    for name in wb_absence.sheetnames:  # cojemos sheets del fichero una a una (2014 a 2020)
        print(name)
        sheet = wb_absence[name]  # cargar la sheet correspondiente
        new_data = [create.create_row_absence(row, header_absence, names_dict, dictionary, df_absence) \
                    for row in islice(sheet.values, 1, sheet.max_row)]
        df_absence = df_absence.append(new_data, ignore_index=True)
    print(df_absence.columns)

    wb_absence.close()
    # print(df_absence.columns)

    df_absence = df_absence.drop(['sociedad', 'numero_de_personal', '%_hortbjo'], axis=1)
    #df_absence['clase_de_absentismo_o_presenci'] = pd.Categorical(df_absence['clase_de_absentismo_o_presenci'])
    #print(df_absence['clase_de_absentismo_o_presenci'].cat.categories)
    #a= input('editar filtro clase absentismos')
    non_absence = ['Asuntos Propios', 'AT Pago directo', 'Cuidado menor enf.grave/c', 'ERE/ERTE (DÍAS)',
                   'ERTE (HORAS)', 'Guarda legal', 'Lactancia (días)', 'Maternidad', 'Paternidad', 'Paternidad/TP',
                   'Permiso no retribuido', 'Vacaciones (d.naturales)', 'Vacaciones internas', 'Vacaciones (d.laborales)']

    #non_absence = ['Guarda legal', 'Vacaciones (d.naturales)', 'Vacaciones internas', 'Vacaciones (d.laborales)']

    df_absence = df_absence[df_absence['clase_de_absentismo_o_presenci'].isin(non_absence) == False]

    """
    .filter("tipo_absentismo not in ('ASUNTOS PROPIOS','AT PAGO DIRECTO','CUIDADO MENOR ENF.GRAVE/C'" +
            ",'ERE/ERTE (DIAS)','ERTE (HORAS)','GUARDA LEGAL','LACTANCIA (DIAS)','MATERNIDAD','PATERNIDAD'" +
            ",'PATERNIDAD/TP','VACACIONES (D.NATURALES)','VACACIONES (D.LABORALES)','VACACIONES INTERNAS')" +
            " and tipo_absentismo is not null"
            );
    """

    KEY = ['division_de_personal', 'no_pers', 'numero_de_orden', 'centro_de_coste', 'desde',
           'identificacion_de_contratos_se', 'hasta', 'dianat']
    df_absence = df_absence.drop_duplicates(KEY, keep='first').sort_values(['no_pers', 'desde'])

    print('\ndf_absence completado en: ', datetime.now() - now, '\n')
    print(df_absence.columns)
    now = datetime.now()

    spark_absenceDF = spark.createDataFrame(df_absence)

    for col in ['desde', 'hasta']:
        spark_absenceDF = spark_absenceDF.withColumn(col, spark_absenceDF[col].cast(DateType()))
    print('Cast de fechas a datetype: ')
    spark_absenceDF.show()
    print('\nPasar de df_absence a sparkdf completado en: ', datetime.now() - now)

    spark_absenceDF = spark_absenceDF.filter(spark_absenceDF.hasta < now) \
                                    .withColumn('numero_de_orden',  when(F.col('numero_de_orden') != "",
                                                                         F.col('numero_de_orden'))
                                                .otherwise(F.lit('Multiservicio')))

    print('Spark absences')
    spark_absenceDF.show()
    spark_absenceDF.toPandas().to_csv('c:\\tmp\\absence.csv')
    #a = input('Absence.csv salvado. Continuar?')

    ########################################################
    ###    JOIN DE WORKERS Y AUSENCIAS Y GENERACIÓN DE   ###
    ###   LAS FECHAS INICIO/FIN DE ACUERDO AL CONTRATO   ###
    ########################################################

    now = datetime.now()

    print('\ndataframes de workers y ausencias completados')

    # cero_days = timedelta(days=0)

    # sparkDF = sparkDF.withColumn("ausencias", lit("0"))
    sparkDF.createOrReplaceTempView("datos_trabajador")
    spark_absenceDF.createOrReplaceTempView("ausencias")

    # df.loc[index, 'ausencia'] += (min(upper_absence, upper_worker) - max(lower_absence, lower_worker) + \
    #                              np.timedelta64(1, 'D'))

    sql = "SELECT a.*, b.desde AS inicio, b.hasta AS fin, b.clase_de_absentismo_o_presenci AS clase_absentismo, " + \
          "CASE WHEN b.desde IS NOT NULL THEN " + \
            "CASE WHEN a.desde_real <= b.desde THEN b.desde ELSE a.desde_real END " + \
          "ELSE b.desde END AS inicial_real, " + \
          "CASE WHEN b.hasta IS NOT NULL THEN " + \
            "CASE WHEN a.hasta_real >= b.hasta THEN b.hasta ELSE a.hasta_real END " + \
          "ELSE b.hasta END AS final_real " + \
          "FROM datos_trabajador AS a LEFT JOIN ausencias as b " + \
          "ON b.division_de_personal = a.division_de_personal and b.no_pers = a.no_pers and " +\
          "b.centro_de_coste = a.centro_de_coste and b.numero_de_orden = a.numero_de_orden and " +\
          "b.identificacion_de_contratos_se = a.identificacion_de_contratos_se and b.desde <= a.hasta and " +\
          "b.hasta>=a.desde"  # or dianat = 0)

    sparkDF = spark.sql(sql)
    print('Join')
    sparkDF.show()
    # a = input('Proseguir?')
    # sparkDF.sort('no_pers', 'desde_real', 'inicial_real').show()
    print('\nJoin sql completado en: ', datetime.now() - now)

    now = datetime.now()

    #######################################
    ### GENERAR COLUMNAS DE DIFERENCIAS ###
    ###  TEMPORALES PARA EL RANGO, LAS  ###
    ###    AUSENCIAS EN DÍAS Y EN %     ###
    #######################################

    sparkDF2 = sparkDF.withColumn("ausencia_abs", when(F.col('inicial_real').isNull(), F.lit(0))
                                                 .otherwise(F.datediff('final_real', 'inicial_real') + 1))\
                      .withColumn("ausencia_rel", F.col('ausencia_abs') / F.col('dias_rango'))\
                      .withColumn('clase_absentismo', when(F.col('clase_absentismo').isNull(), F.lit('ninguna'))
                                                     .otherwise(F.col('clase_absentismo')))

    #print('join.csv generado con las cols inicial_real y final_real')
    #sparkDF2.toPandas().to_csv('c:\\tmp\\join.csv')


    ##############################################
    ### AGRUPAR TODAS LAS AUSENCIAS PARA TENER ###
    ###  EL TOTAL Y DIVIDIR POR EL TIEMPO DE   ###
    ###  CONTRATO PARA TENER EL % FALTANTE     ###
    ##############################################

    # sparkDF2 = sparkDF.drop('year', 'hasta_max', 'inicio', 'fin', 'desde', 'hasta')
    # sparkDF2.sort('no_pers', 'desde_real').show()
    # sparkDF2.printSchema()
    sparkDF2.createOrReplaceTempView("datos_trabajador")

    KEY = ['division_de_personal', 'no_pers', 'clave_de_sexo', 'descripcion', 'porcentaje_de_discapacidad',
           'numero_de_orden', 'centro_de_coste', 'categoria_profesional', 'clase_de_contrato',
           'identificacion_de_contratos_se', 'h_sem', 'feantig', 'fecha_nac', 'nacionalidad',
           'codpostal', 'clase_de_instituto_de_ensenanz', 'clase_registro_parentesco', 'idobj', 'desde_real',
           'hasta_real', 'dias_rango']

    sparkDF_agg = sparkDF2.groupBy(KEY).sum('ausencia_abs') \
        .withColumnRenamed("sum(ausencia_abs)", "ausencia_total_abs") \
        .withColumn('ausencia_total_rel', F.col('ausencia_total_abs') / F.col('dias_rango'))
    sparkDF_agg.groupBy().agg(F.sum(F.col('ausencia_total_abs')), F.sum(F.col('dias_rango'))).show()
    #print('Ausencias y total: ', counts)
    # sparkDF_agg.printSchema()
    # sparkDF_agg.sort('no_pers', 'desde_real').show()

    # sparkDF_agg.toPandas().to_csv('c:\\tmp\\agg.csv')

    sparkDF_agg.createOrReplaceTempView("ausencias")

    sql = "SELECT a.*, b.ausencia_total_abs, b.ausencia_total_rel FROM datos_trabajador AS a LEFT JOIN ausencias as b " \
          "ON b.no_pers = a.no_pers"
    KEY.remove('no_pers')
    adition = ""
    for element in KEY:
        adition += " and b." + element + " = a." + element
    sql += adition
    print(sql)
    sparkDF2 = spark.sql(sql)
    sparkDF2.printSchema()
    # sparkDF2.select(F.col('no_pers'), F.col('desde_real'), F.col('hasta_real'), F.col('dias_rango'), F.col('inicial_real'),
    #                F.col('final_real'), F.col('ausencia_abs'), F.col('ausencia_total_abs'), F.col('ausencia_total_rel'))\
    #                .sort('no_pers', 'desde_real', 'inicial_real').show()

    ####################################
    ###   ARREGLAR LOS ERRORES CON   ###
    ###   LAS FECHAS DE ANTIGUEDAD   ###
    ####################################

    sparkDF2 = sparkDF2.withColumn('feantig', when(F.col('feantig') > F.col('desde_real'), F.col('desde_real'))
                                   .otherwise(F.col('feantig')))

    #############################
    ### SALVAR EN MONGO Y CSV ###
    #############################

    #sparkDF2.printSchema()
    #sparkDF2.show()
    # sparkDF2.select([F.count(when(F.col(c) == "")).alias(c) for c in sparkDF2.columns]).show()


    sparkDF2.write.format("mongo").mode("overwrite").save()

    print('\nEscritura a mongo completada en: ', datetime.now() - now)
    # docNormalized.printSchema()

    now = datetime.now()

    # sparkDF2.toPandas().to_csv('c:\\tmp\\final.csv')

    # print('\nEscritura a csv completada en: ', datetime.now() - now)

    spark.stop()

    return True